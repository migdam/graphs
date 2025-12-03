#!/usr/bin/env python3
"""
Embed Charts in Excel Files
Generate charts and embed them directly into Excel workbooks.
"""

import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import os
import sys
import argparse
import logging
from io import BytesIO
from pathlib import Path

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)

# Add scripts directory to path
sys.path.insert(0, os.path.dirname(__file__))
from excel_utils import read_excel


def create_chart_image(df, chart_type='line', **kwargs):
    """
    Create a chart and return as image bytes.
    
    Args:
        df: DataFrame with data
        chart_type: Type of chart to create
        **kwargs: Additional chart parameters
        
    Returns:
        BytesIO object containing PNG image
    """
    fig, ax = plt.subplots(figsize=(8, 5))
    
    try:
        if chart_type == 'line':
            for category in df['category'].unique() if 'category' in df.columns else [None]:
                if category:
                    subset = df[df['category'] == category]
                    ax.plot(subset.index, subset['value'], marker='o', label=category)
                else:
                    ax.plot(df.index, df['value'], marker='o')
            ax.set_title('Line Chart', fontweight='bold', fontsize=12)
            ax.legend()
            
        elif chart_type == 'bar':
            if 'category' in df.columns:
                data = df.groupby('category')['value'].sum()
                data.plot(kind='bar', ax=ax, color=sns.color_palette('viridis', len(data)))
            else:
                df['value'].plot(kind='bar', ax=ax)
            ax.set_title('Bar Chart', fontweight='bold', fontsize=12)
            
        elif chart_type == 'pie':
            if 'category' in df.columns:
                data = df.groupby('category')['value'].sum()
                ax.pie(data, labels=data.index, autopct='%1.1f%%', startangle=90)
            else:
                df['value'].plot(kind='pie', ax=ax, autopct='%1.1f%%')
            ax.set_title('Distribution', fontweight='bold', fontsize=12)
            
        elif chart_type == 'scatter':
            x_col = kwargs.get('x', df.columns[0])
            y_col = kwargs.get('y', df.columns[1] if len(df.columns) > 1 else df.columns[0])
            ax.scatter(df[x_col], df[y_col], alpha=0.6, s=100)
            ax.set_xlabel(x_col)
            ax.set_ylabel(y_col)
            ax.set_title('Scatter Plot', fontweight='bold', fontsize=12)
        
        ax.grid(True, alpha=0.3)
        plt.tight_layout()
        
        # Save to BytesIO
        img_buffer = BytesIO()
        fig.savefig(img_buffer, format='png', dpi=100, bbox_inches='tight')
        img_buffer.seek(0)
        
        plt.close(fig)
        return img_buffer
        
    except Exception as e:
        logger.error(f"Error creating chart: {str(e)}")
        plt.close(fig)
        return None


def embed_chart_in_excel(excel_path, output_path, sheet_name=None, 
                         chart_type='line', position='H2', **kwargs):
    """
    Embed a chart into an Excel file.
    
    Args:
        excel_path: Path to input Excel file
        output_path: Path to save output Excel file
        sheet_name: Sheet to read data from and embed chart into
        chart_type: Type of chart to create
        position: Cell position to place chart (e.g., 'H2')
        **kwargs: Additional chart parameters
        
    Returns:
        True if successful, False otherwise
    """
    try:
        # Read data from Excel
        logger.info(f"Reading data from {excel_path}")
        df = read_excel(excel_path, sheet_name=sheet_name)
        
        if df is None:
            logger.error("Could not read Excel file")
            return False
        
        # Create chart image
        logger.info(f"Creating {chart_type} chart")
        img_buffer = create_chart_image(df, chart_type=chart_type, **kwargs)
        
        if img_buffer is None:
            return False
        
        # Load workbook
        logger.info("Embedding chart into Excel")
        
        # If output is same as input, work with original
        # Otherwise, copy first
        if excel_path != output_path:
            import shutil
            shutil.copy2(excel_path, output_path)
        
        wb = load_workbook(output_path)
        
        # Select sheet
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        # Create image from buffer
        img = Image(img_buffer)
        
        # Resize image to fit nicely in Excel
        img.width = 500
        img.height = 312
        
        # Add image to worksheet
        ws.add_image(img, position)
        
        # Save workbook
        wb.save(output_path)
        logger.info(f"Chart embedded successfully in {output_path}")
        
        return True
        
    except Exception as e:
        logger.error(f"Error embedding chart: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def embed_multiple_charts(excel_path, output_path, charts_config):
    """
    Embed multiple charts into an Excel file.
    
    Args:
        excel_path: Path to input Excel file
        output_path: Path to save output Excel file
        charts_config: List of chart configurations
            Each config is a dict with: sheet, chart_type, position
            
    Example:
        charts_config = [
            {'sheet': 'Sales', 'chart_type': 'line', 'position': 'H2'},
            {'sheet': 'Sales', 'chart_type': 'bar', 'position': 'H20'},
            {'sheet': 'Costs', 'chart_type': 'pie', 'position': 'H2'},
        ]
    
    Returns:
        True if successful, False otherwise
    """
    try:
        import shutil
        
        # Copy input to output
        if excel_path != output_path:
            shutil.copy2(excel_path, output_path)
        
        # Process each chart
        for i, config in enumerate(charts_config, 1):
            logger.info(f"Processing chart {i}/{len(charts_config)}")
            
            sheet = config.get('sheet')
            chart_type = config.get('chart_type', 'line')
            position = config.get('position', 'H2')
            
            # Read data
            df = read_excel(output_path, sheet_name=sheet)
            if df is None:
                logger.warning(f"Skipping chart {i}: Could not read sheet '{sheet}'")
                continue
            
            # Create chart
            img_buffer = create_chart_image(df, chart_type=chart_type)
            if img_buffer is None:
                logger.warning(f"Skipping chart {i}: Could not create chart")
                continue
            
            # Embed chart
            wb = load_workbook(output_path)
            ws = wb[sheet] if sheet in wb.sheetnames else wb.active
            
            img = Image(img_buffer)
            img.width = 500
            img.height = 312
            
            ws.add_image(img, position)
            wb.save(output_path)
        
        logger.info(f"All charts embedded successfully in {output_path}")
        return True
        
    except Exception as e:
        logger.error(f"Error embedding multiple charts: {str(e)}")
        return False


def create_report_with_charts(data_dict, output_path, chart_positions=None):
    """
    Create an Excel report with data and embedded charts.
    
    Args:
        data_dict: Dictionary of {sheet_name: DataFrame}
        output_path: Path to save Excel file
        chart_positions: Dict of {sheet_name: {'chart_type': ..., 'position': ...}}
        
    Returns:
        True if successful, False otherwise
    """
    try:
        # Write data to Excel
        logger.info(f"Creating report: {output_path}")
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for sheet_name, df in data_dict.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Embed charts if specified
        if chart_positions:
            wb = load_workbook(output_path)
            
            for sheet_name, chart_config in chart_positions.items():
                if sheet_name not in data_dict:
                    continue
                
                df = data_dict[sheet_name]
                chart_type = chart_config.get('chart_type', 'line')
                position = chart_config.get('position', 'H2')
                
                # Create chart
                img_buffer = create_chart_image(df, chart_type=chart_type)
                if img_buffer:
                    ws = wb[sheet_name]
                    img = Image(img_buffer)
                    img.width = 500
                    img.height = 312
                    ws.add_image(img, position)
            
            wb.save(output_path)
        
        logger.info(f"Report created successfully: {output_path}")
        return True
        
    except Exception as e:
        logger.error(f"Error creating report: {str(e)}")
        return False


def main():
    parser = argparse.ArgumentParser(
        description='Embed charts into Excel files',
        formatter_class=argparse.ArgumentDefaultsHelpFormatter
    )
    
    parser.add_argument('excel_file', help='Input Excel file')
    parser.add_argument('--output', help='Output Excel file (default: adds _charts suffix)')
    parser.add_argument('--sheet', help='Sheet name to use (default: first sheet)')
    parser.add_argument('--chart-type', default='line',
                       choices=['line', 'bar', 'pie', 'scatter'],
                       help='Type of chart to create')
    parser.add_argument('--position', default='H2',
                       help='Cell position for chart (e.g., H2, M10)')
    parser.add_argument('--x', help='X column for scatter plot')
    parser.add_argument('--y', help='Y column for scatter plot')
    
    args = parser.parse_args()
    
    # Determine output path
    if args.output:
        output_path = args.output
    else:
        base = Path(args.excel_file)
        output_path = str(base.parent / f"{base.stem}_with_chart{base.suffix}")
    
    # Embed chart
    kwargs = {}
    if args.x:
        kwargs['x'] = args.x
    if args.y:
        kwargs['y'] = args.y
    
    success = embed_chart_in_excel(
        args.excel_file,
        output_path,
        sheet_name=args.sheet,
        chart_type=args.chart_type,
        position=args.position,
        **kwargs
    )
    
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
